from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment

import os
import psycopg2
import re

databases = [('al00scm', 'hmj40bfac01'), ('kb00scm', 'hmj40bfac01'),
             ('km00scm', 'hmj40bfac01'), ('ks00scm', 'hmj40bfac01'),
             ('ks00scm', 'hmj40bfac02'), ('mo00scm', 'hmj40bfam01')]

SCHEMA_LIST = ['km00scm', 'ks00scm', 'mo00scm']

trigger_result_data = [
    ['object_type', 'database_name', 'schema_name', 'table_name', 'trigger_name', 'description', 'event',
     'trigger_body', 'note']
]

function_result_data = [
    ['object_type', 'database_name', 'schema_name', 'function_name', 'function_body']
]

procedure_result_data = [
    ['object_type', 'database_name', 'schema_name', 'procedure_name', 'procedure_body', 'note']
]

view_table_result_data = [
    ['object_type', 'database_name', 'schema_name', 'view_name', 'view_body', 'note']
]

columns_result_data = [
    ['object_type', 'database_name', 'schema_name',
     'table_name', 'column_name', 'data_type', 'constraint_name', 'constraint_type', 'note']
]

update_result_data = [
    ['object_type', 'object_name', 'update_object_type', 'update_object_name']
]

object_result_data = [
    ['database_name', 'schema_name', 'object_type', 'object_name', 'machine_id']
]

trigger_syntax = '''
    SELECT tgname, pg_get_triggerdef(tr.oid) AS trigger_def, t.relname, n.nspname, pg_get_functiondef(tr.tgfoid) AS function_definition, *
    FROM pg_trigger tr
    INNER JOIN pg_class AS t ON tr.tgrelid = t.oid
    INNER JOIN pg_namespace n ON n.oid = t.relnamespace
    '''

function_syntax = '''
    SELECT p.proname, nspname, p.prosrc, t.typname, *
    FROM pg_catalog.pg_proc p
    JOIN pg_catalog.pg_namespace n ON n.oid = p.pronamespace
    JOIN pg_catalog.pg_language ON p.prolang = pg_catalog.pg_language.oid
    JOIN pg_type t ON p.prorettype = t.oid
    where pg_catalog.pg_language.lanname = 'plpgsql';
'''

view_table_syntax = '''
    SELECT schemaname, viewname, definition
    FROM pg_views
    WHERE schemaname <> 'information_schema' and schemaname <> 'pg_catalog'
'''

column_list_syntax = '''
    SELECT cols.table_schema, cols.table_name, cols.column_name, cols.data_type, col_cons.constraint_name,  cons.constraint_type
    FROM information_schema.columns cols
    LEFT JOIN information_schema.constraint_column_usage col_cons ON cols.table_schema = col_cons.table_schema
                                                               AND cols.table_name = col_cons.table_name
                                                               AND cols.column_name = col_cons.column_name
    LEFT JOIN information_schema.table_constraints cons ON col_cons.constraint_name = cons.constraint_name
                                                         AND col_cons.constraint_schema = cons.constraint_schema
    WHERE cols.table_schema <> 'pg_catalog'
'''

table_list_syntax = '''
    SELECT table_name, table_schema
    FROM information_schema.tables
    WHERE table_schema <> 'postgres' and table_schema <> 'information_schema' and table_schema <> 'pg_catalog';
'''

def extract_words_before_pattern(sentence):
    match = re.search(r'.+' + r'(BEFORE \w+)\b', sentence)
    if match:
        return match.group(1)

    return None

def extract_word_after_word_update(sentence):
    match = re.search(r'.+' + r'update (\w+)\b', sentence)
    if match:
        return match.group(1)

    return ''

def handle_object_type(object_name):
    if object_name in SCHEMA_LIST:
        return 'SCHEMA'

    return 'TABLE'


def list_trigger(database_name):
    conn = connect_to_db(database_name)
    cur = conn.cursor()

    cur.execute(trigger_syntax)
    results = cur.fetchall()

    trigger_name_index = 0
    trigger_def_index = 1
    trigger_table_name_index = 2
    trigger_schema_name_index = 3
    trigger_func_def_index = 4
    for row in results:
        trigger_data = ['TRIGGER', database_name, row[trigger_schema_name_index], row[trigger_table_name_index],
                        row[trigger_name_index], row[trigger_def_index],
                        extract_words_before_pattern(row[trigger_def_index]),
                        row[trigger_func_def_index].strip(), '']
        trigger_result_data.append(trigger_data)
        trigger_object_name = extract_word_after_word_update(row[trigger_func_def_index])
        if trigger_object_name != '':
            trigger_object_type = handle_object_type(trigger_object_name)
            trigger_object_data = ['TRIGGER', row[trigger_name_index], trigger_object_type, trigger_object_name]
            update_result_data.append(trigger_object_data)

    cur.close()
    conn.close()

def list_view_tables(database_name):
    conn = connect_to_db(database_name)
    cur = conn.cursor()

    cur.execute(view_table_syntax)
    results = cur.fetchall()

    view_table_def_index = 2
    view_table_schema_name = 0
    view_table_name_index = 1
    for row in results:
        temp = ['VIEW', database_name, row[view_table_schema_name], row[view_table_name_index],
                row[view_table_def_index].strip(), '']
        view_table_result_data.append(temp)

    cur.close()
    conn.close()

def list_functions(database_name):
    conn = connect_to_db(database_name)
    cur = conn.cursor()

    cur.execute(function_syntax)
    results = cur.fetchall()

    function_name_index = 0
    function_def_index = 2
    function_ret_type_index = 3
    function_schema_name_index = 1

    for row in results:
        temp = ['FUNCTION', database_name, row[function_schema_name_index],
                row[function_name_index], row[function_def_index].strip()]

        if row[function_ret_type_index].lower() == 'void':
            temp[0] = 'PROCEDURE'
            procedure_result_data.append(temp)
        else:
            function_result_data.append(temp)

    cur.close()
    conn.close()

def list_columns(database_name):
    conn = connect_to_db(database_name)
    cur = conn.cursor()

    cur.execute(column_list_syntax)
    results = cur.fetchall()

    column_name_index = 2
    column_schema_name_index = 0
    column_table_name_index = 1
    column_type_index = 3
    column_constraint_name_index = 4
    column_constraint_type_index = 5

    for row in results:
        temp = ['TABLE', database_name, row[column_schema_name_index], row[column_table_name_index],
                row[column_name_index], row[column_type_index], row[column_constraint_name_index],
                row[column_constraint_type_index], '']
        columns_result_data.append(temp)

    cur.close()
    conn.close()

def list_object(database_name):
    conn = connect_to_db(database_name)
    cur = conn.cursor()

    cur.execute(table_list_syntax)
    results = cur.fetchall()

    table_schema_name_index = 1
    table_name_index = 0

    for row in results:
        temp = [database_name, row[table_schema_name_index], 'TABLE', row[table_name_index], '']
        object_result_data.append(temp)

    function_name_index = 3
    function_schema_name_index = 2

    for row in function_result_data:
        temp = [database_name, row[function_schema_name_index], 'FUNCTION', row[function_name_index], '']
        object_result_data.append(temp)

    procedure_name_index = 3
    procedure_schema_name_index = 2

    for row in procedure_result_data:
        temp = [database_name, row[procedure_schema_name_index], 'PROCEDURE', row[procedure_name_index], '']
        object_result_data.append(temp)

    trigger_name_index = 4
    trigger_schema_name_index = 2

    for row in trigger_result_data:
        temp = [database_name, row[trigger_schema_name_index], 'TRIGGER', row[trigger_name_index], '']
        object_result_data.append(temp)

    view_table_name_index = 3
    view_table_schema_name_index = 2

    for row in view_table_result_data:
        temp = [database_name, row[view_table_schema_name_index], 'VIEW', row[view_table_name_index], '']
        object_result_data.append(temp)

    cur.close()
    conn.close()

def export_data_to_sheet(wb, sheet_name, input_data, c_height_pivot_idx):
    result_sheet = None
    if sheet_name in wb.sheetnames:
        result_sheet = wb[sheet_name]
    else:
        result_sheet = wb.create_sheet(title=sheet_name)

    for row_idx, row_data in enumerate(input_data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = result_sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            if result_sheet.cell(row=row_idx, column=col_idx).alignment:
                cell.alignment = Alignment(**result_sheet.cell(row=row_idx, column=col_idx).alignment.__dict__)

        if c_height_pivot_idx != -1:
            height_calc = 14.4 * (row_data[c_height_pivot_idx].count("\n") + 1)
            cell_height = 409 if height_calc > 409 else height_calc
            result_sheet.row_dimensions[row_idx].height = cell_height

def connect_to_db(database_name):
    conn = psycopg2.connect(
        dbname=database_name,
        user="postgres",
        password="1234",
        host="localhost",
        port="5432",
    )
    return conn


for database in databases:
    database_name = database[1]
    list_columns(database_name)
    list_view_tables(database_name)
    list_trigger(database_name)
    list_functions(database_name)
    list_object(database_name)

filename = "result3.xlsx"

ws = None
if os.path.exists(filename):
    wb = load_workbook(filename)
else:
    wb = Workbook()

#export_data_to_sheet(wb, 'table', columns_result_data, -1)
#export_data_to_sheet(wb, 'view', view_table_result_data, 4)
#export_data_to_sheet(wb, 'trigger', trigger_result_data, 7)
#export_data_to_sheet(wb, 'procedure', procedure_result_data, 4)
#export_data_to_sheet(wb, 'function', function_result_data, 4)
#export_data_to_sheet(wb, 'update', update_result_data, -1)
export_data_to_sheet(wb, 'object', object_result_data, -1)


wb.save(filename)
